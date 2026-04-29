"""
Contact Export — Lista pronta para contacto direto.

Gera CSV e XLSX com leads contactáveis (telefone válido),
deduplicados por telefone, ordenados por score desc.
Pronto para envio ao cliente sem tratamento manual.

Uso:
    python main.py export-contacts
    python main.py export-contacts --score-min 40 --zones Lisboa,Cascais
    python main.py export-contacts --format xlsx
    python main.py export-contacts --include-agencies
"""
from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from config.settings import settings
from storage.database import get_db
from storage.models import Lead
from utils.logger import get_logger

log = get_logger(__name__)

# ── Urgency keyword detection (for insight column) ────────────────────────────
_URGENCY_KW: list[tuple[str, str]] = [
    (r"\burgente\b|\burgência\b",             "venda urgente"),
    (r"\bherança\b|\bherdeiro\b|\bpartilha\b", "imóvel de herança"),
    (r"\bdivórcio\b|\bseparação\b",            "processo de divórcio"),
    (r"\bemigr\w+\b",                          "proprietário a emigrar"),
    (r"\bexecutiv\w+\b|\bpenhora\b",           "execução hipotecária"),
    (r"\bpermuta\b",                           "aberto a permuta"),
    (r"\bnegocia\w*\b",                        "preço negociável"),
    (r"\b(sem comissão|sem mediadora|direto do dono|proprietário vende)\b",
     "direto do proprietário"),
    (r"\bpreciso vender\b|\bprecisa vender\b", "precisa vender"),
    (r"\bpara recuperar\b|\bpara remodelar\b", "imóvel para remodelar"),
]
_URGENCY_COMPILED = [(re.compile(p, re.IGNORECASE), label) for p, label in _URGENCY_KW]


# ── Human-readable type labels ─────────────────────────────────────────────────
_LEAD_TYPE_LABELS = {
    "fsbo":              "Proprietário Venda",
    "frbo":              "Proprietário Arrendamento",
    "agency_listing":    "Agência",
    "developer_listing": "Promotor",
    "active_owner":      "Potencial Proprietário",
    "unknown":           "Particular/Desconhecido",
}

_OWNER_TYPE_LABELS = {
    "fsbo":      "Proprietário",
    "agency":    "Agência",
    "developer": "Promotor",
    "unknown":   "Desconhecido",
}

_SOURCE_LABELS = {
    "olx":             "OLX",
    "imovirtual":      "Imovirtual",
    "idealista":       "Idealista",
    "sapo":            "Sapo Casa",
    "custojusto":      "CustoJusto",
    "olx_marketplace": "OLX Marketplace",
    "standvirtual":    "StandVirtual",
    "linkedin":        "LinkedIn",
}

_SCORE_LABELS = {
    "HOT":  "🔴 HOT",
    "WARM": "🟡 WARM",
    "COLD": "🔵 COLD",
}

def _canonical_phone(phone: str) -> Optional[str]:
    """Return canonical +351XXXXXXXXX or None if invalid/suspicious."""
    from utils.phone import validate_pt_phone
    result = validate_pt_phone(phone or "")
    return result.canonical if result.valid else None


def _phone_type_label(phone: str) -> str:
    """Return human-readable phone type for the export."""
    from utils.phone import validate_pt_phone
    result = validate_pt_phone(phone or "")
    if not result.valid:
        return "inválido"
    labels = {
        "mobile":   "Telemóvel",
        "landline": "Fixo",
        "relay":    "Relay/OLX",
        "unknown":  "Desconhecido",
    }
    return labels.get(result.phone_type, result.phone_type)


def _whatsapp_link(phone: str) -> str:
    """Build wa.me link with pre-filled greeting message."""
    digits = phone.replace("+", "").replace(" ", "")
    msg = "Olá, vi o seu anúncio e gostaria de saber mais sobre o imóvel. Pode falar?"
    from urllib.parse import quote
    return f"https://wa.me/{digits}?text={quote(msg)}"


def _get_url(sources_json: str) -> str:
    """Extract first listing URL from sources JSON."""
    try:
        sources = json.loads(sources_json or "[]")
        if sources:
            return sources[0].get("url", "")
    except (json.JSONDecodeError, IndexError, KeyError):
        pass
    return ""


def _build_insight(lead: Lead) -> str:
    """
    Generate a short actionable insight for this lead.
    Combines urgency signals, days on market, and price delta.
    """
    parts: list[str] = []

    # Urgency keywords from description + title
    text = f"{lead.title or ''} {lead.description or ''}"
    for pattern, label in _URGENCY_COMPILED:
        if pattern.search(text):
            parts.append(label)
            break  # one urgency signal is enough

    # Days on market
    dom = lead.days_on_market or 0
    if dom >= 90:
        parts.append(f"{dom} dias no mercado")
    elif dom >= 30:
        parts.append(f"{dom} dias listado")

    # Price delta
    delta = lead.price_delta_pct
    if delta and delta >= 15:
        parts.append(f"{delta:.0f}% abaixo do mercado")
    elif delta and delta >= 5:
        parts.append(f"{delta:.0f}% abaixo do benchmark")

    # Owner type
    if lead.owner_type == "fsbo" and not parts:
        parts.append("contacto direto com proprietário")
    elif lead.lead_type == "frbo" and not parts:
        parts.append("senhorio a arrendar diretamente")

    # Mobile phone bonus
    phone = lead.contact_phone or ""
    national = phone.replace("+351", "")
    if national.startswith("9") and not any("direto" in p for p in parts):
        parts.append("telemóvel direto")

    return "; ".join(parts) if parts else "—"


def _is_likely_agency(lead: Lead) -> bool:
    """Return True when this lead is likely an agency (not a direct owner)."""
    if lead.owner_type == "agency":
        return True
    if lead.lead_type == "agency_listing":
        return True
    if not lead.is_owner and lead.agency_name:
        return True
    # Landline 21x/22x = very likely agency
    phone = (lead.contact_phone or "").replace("+351", "")
    if phone.startswith("21") or phone.startswith("22"):
        return True
    return False


def _format_price(price: Optional[float]) -> str:
    if not price:
        return "—"
    return f"€ {price:,.0f}".replace(",", ".")


_ROOM_RENTAL_RE = re.compile(
    r"\bquarto\b|\bquartos\b|\bcama\b|\bcamas\b|\bvaga\s+em\b|\bquarto\s+partilhado\b",
    re.IGNORECASE,
)


def _confidence_label(lead: Lead) -> str:
    """
    Return a confidence indicator for the owner classification.

    HIGH  — explicit FSBO/FRBO signal in text, or scraper-confirmed owner
    MED   — source bias (OLX/custojusto) or has mobile number
    LOW   — unknown classification, landline only
    """
    if lead.owner_type == "fsbo" and lead.lead_type in ("fsbo", "frbo"):
        phone = (lead.contact_phone or "").replace("+351", "")
        if phone.startswith("9"):
            return "ALTA"
        return "MÉDIA"
    if lead.owner_type in ("unknown",):
        return "BAIXA"
    return "MÉDIA"


def generate_contact_list(
    score_min: int = 0,
    zones: Optional[list[str]] = None,
    include_agencies: bool = False,
    exclude_room_rentals: bool = True,
    mobile_only: bool = False,
    limit: int = 2000,
) -> list[dict]:
    """
    Build the contact list from the database.

    Returns a list of dicts, one per unique phone number,
    sorted by score descending.
    """
    from sqlalchemy import select, and_, or_

    with get_db() as db:
        q = select(Lead).where(
            Lead.archived == False,                           # noqa: E712
            Lead.is_demo == False,                            # noqa: E712
            Lead.contact_phone.isnot(None),
            Lead.contact_phone != "",
            Lead.score >= score_min,
        )
        if zones:
            q = q.where(Lead.zone.in_(zones))
        if not include_agencies:
            q = q.where(
                or_(
                    Lead.owner_type != "agency",
                    Lead.owner_type.is_(None),
                )
            )
        q = q.order_by(Lead.score.desc()).limit(limit * 3)  # extra buffer for dedup
        leads = db.execute(q).scalars().all()

    # Deduplicate by canonical phone — keep highest score per phone
    seen_phones: dict[str, dict] = {}
    rows: list[dict] = []

    for lead in leads:
        phone = _canonical_phone(lead.contact_phone or "")
        if not phone:
            continue  # invalid format — skip

        if phone in seen_phones:
            continue  # already have this contact at higher score (ordered desc)

        # Skip room rentals — not target property type for investor leads
        title_str = lead.title or ""
        if exclude_room_rentals and _ROOM_RENTAL_RE.search(title_str):
            continue

        # Classify phone type (use stored field if available, else compute)
        stored_type = getattr(lead, "phone_type", None) or ""
        tipo_tel = _phone_type_label(phone) if not stored_type or stored_type == "unknown" else {
            "mobile":   "Telemóvel",
            "landline": "Fixo",
            "relay":    "Relay/OLX",
        }.get(stored_type, stored_type)

        # Optional filter: only mobile numbers
        if mobile_only and stored_type not in ("mobile", ""):
            # also check by phone prefix when stored_type not set
            from utils.phone import validate_pt_phone
            _r = validate_pt_phone(phone)
            if _r.phone_type != "mobile":
                continue

        is_agency = _is_likely_agency(lead)
        url = _get_url(lead.sources_json or "[]")

        row = {
            "score":          lead.score,
            "label":          _SCORE_LABELS.get(lead.score_label, lead.score_label),
            "nome":           lead.contact_name or "—",
            "telefone":       phone,
            "tipo_telefone":  tipo_tel,
            "whatsapp":       _whatsapp_link(phone),
            "zona":           lead.zone or "—",
            "concelho":       lead.municipality or lead.zone or "—",
            "tipologia":      lead.typology or "—",
            "preco":          _format_price(lead.price),
            "area_m2":        f"{lead.area_m2:.0f} m²" if lead.area_m2 else "—",
            "tipo_lead":      _LEAD_TYPE_LABELS.get(lead.lead_type or "unknown", lead.lead_type or "—"),
            "fonte":          _SOURCE_LABELS.get(lead.discovery_source or "", lead.discovery_source or "—"),
            "agencia":        "SIM" if is_agency else "não",
            "confianca":      _confidence_label(lead),
            "insight":        _build_insight(lead),
            "url":            url,
            "titulo":         lead.title or "—",
            "data":           lead.first_seen_at.strftime("%d/%m/%Y") if lead.first_seen_at else "—",
            "dias_mercado":   str(lead.days_on_market or 0),
        }

        seen_phones[phone] = row
        rows.append(row)

        if len(rows) >= limit:
            break

    log.info(
        "Contact list built: {n} leads with phone (from {total} candidates, {dupes} dupes removed)",
        n=len(rows),
        total=len(leads),
        dupes=len(leads) - len(rows),
    )
    return rows


def export_contact_csv(
    rows: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """Export contact list to CSV. Returns file path."""
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(settings.data_dir / f"contactos_{ts}.csv")

    fields = [
        "score", "label", "nome", "telefone", "tipo_telefone", "whatsapp",
        "zona", "concelho", "tipologia", "preco", "area_m2",
        "tipo_lead", "fonte", "agencia", "confianca", "insight",
        "titulo", "url", "data", "dias_mercado",
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig for Excel
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    log.info("CSV exported → {p} ({n} rows)", p=output_path, n=len(rows))
    return output_path


def export_contact_xlsx(
    rows: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """Export contact list to XLSX with formatting. Returns file path."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Install with: pip install openpyxl"
        )

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(settings.data_dir / f"contactos_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos"

    # ── Colour palette ────────────────────────────────────────────────────────
    COL_HEADER_BG  = "1F3864"   # dark blue
    COL_HEADER_FG  = "FFFFFF"
    COL_HOT_BG     = "FFE0E0"   # light red
    COL_WARM_BG    = "FFF8E0"   # light yellow
    COL_AGENCY_BG  = "F0F0F0"   # light grey
    COL_ALT_BG     = "F8FBFF"   # very light blue for alternating rows
    COL_BORDER     = "D0D7E0"

    thin_border = Border(
        left=Side(style="thin", color=COL_BORDER),
        right=Side(style="thin", color=COL_BORDER),
        top=Side(style="thin", color=COL_BORDER),
        bottom=Side(style="thin", color=COL_BORDER),
    )

    # ── Column definitions ────────────────────────────────────────────────────
    columns = [
        ("Score",      "score",        7),
        ("Label",      "label",        9),
        ("Nome",       "nome",         22),
        ("Telefone",   "telefone",     16),
        ("Tipo Tel.",  "tipo_telefone", 12),
        ("WhatsApp",   "whatsapp",     35),
        ("Zona",       "zona",         12),
        ("Concelho",   "concelho",     14),
        ("Tipologia",  "tipologia",    11),
        ("Preço",      "preco",        14),
        ("Área",       "area_m2",      10),
        ("Tipo",       "tipo_lead",    24),
        ("Fonte",      "fonte",        14),
        ("Agência?",   "agencia",      9),
        ("Confiança",  "confianca",    10),
        ("Insight",    "insight",      38),
        ("Título",     "titulo",       45),
        ("URL",        "url",          50),
        ("Data",       "data",         12),
        ("Dias Merc.", "dias_mercado", 10),
    ]

    # ── Header row ────────────────────────────────────────────────────────────
    header_font  = Font(name="Calibri", bold=True, color=COL_HEADER_FG, size=10)
    header_fill  = PatternFill("solid", fgColor=COL_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = thin_border
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_font       = Font(name="Calibri", size=9)
    link_font       = Font(name="Calibri", size=9, color="1F78B4", underline="single")
    center_align    = Alignment(horizontal="center", vertical="center")
    left_align      = Alignment(horizontal="left", vertical="center", wrap_text=False)
    number_align    = Alignment(horizontal="right", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        label_raw   = row.get("label", "")
        is_agency   = row.get("agencia", "não") == "SIM"
        score_val   = row.get("score", 0)

        # Row background
        if is_agency:
            bg = COL_AGENCY_BG
        elif "HOT" in label_raw:
            bg = COL_HOT_BG
        elif "WARM" in label_raw:
            bg = COL_WARM_BG
        elif row_idx % 2 == 0:
            bg = COL_ALT_BG
        else:
            bg = "FFFFFF"

        row_fill = PatternFill("solid", fgColor=bg)

        for col_idx, (_, field_key, _) in enumerate(columns, start=1):
            value  = row.get(field_key, "")
            cell   = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill   = row_fill

            # Special formatting
            if field_key == "score":
                cell.font      = Font(name="Calibri", size=9, bold=True)
                cell.alignment = number_align
            elif field_key == "whatsapp":
                # Clickable hyperlink
                if value and value.startswith("http"):
                    cell.hyperlink = value
                    cell.value = "📲 Abrir WhatsApp"
                    cell.font  = link_font
                cell.alignment = center_align
            elif field_key == "url":
                if value and value.startswith("http"):
                    cell.hyperlink = value
                    cell.value = "🔗 Ver Anúncio"
                    cell.font  = link_font
                cell.alignment = center_align
            elif field_key == "agencia":
                cell.font      = Font(name="Calibri", size=9,
                                      bold=is_agency,
                                      color="C00000" if is_agency else "006100")
                cell.alignment = center_align
            elif field_key in ("score", "dias_mercado"):
                cell.alignment = number_align
            elif field_key == "label":
                cell.alignment = center_align
                cell.font      = Font(name="Calibri", size=9, bold=True)
            elif field_key == "insight":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font      = Font(name="Calibri", size=9, italic=True)
            else:
                cell.font      = data_font
                cell.alignment = left_align

        ws.row_dimensions[row_idx].height = 16

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumo")
    hot_count  = sum(1 for r in rows if "HOT"  in r.get("label", ""))
    warm_count = sum(1 for r in rows if "WARM" in r.get("label", ""))
    cold_count = sum(1 for r in rows if "COLD" in r.get("label", ""))
    ag_count   = sum(1 for r in rows if r.get("agencia") == "SIM")
    owner_count = len(rows) - ag_count

    summary_data = [
        ("Relatório de Leads", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Total contactos únicos",  len(rows)),
        ("  🔴 HOT (score ≥ 75)",   hot_count),
        ("  🟡 WARM (score 50-74)", warm_count),
        ("  🔵 COLD (score < 50)",  cold_count),
        ("", ""),
        ("Proprietários directos", owner_count),
        ("Prováveis agências",      ag_count),
        ("", ""),
        ("Zonas cobertas", len({r.get("zona") for r in rows})),
        ("Fontes",         len({r.get("fonte") for r in rows})),
    ]

    title_font = Font(name="Calibri", bold=True, size=12, color=COL_HEADER_BG)
    label_font = Font(name="Calibri", size=10)
    value_font = Font(name="Calibri", bold=True, size=10)

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 22

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        ca = ws_sum.cell(row=r_idx, column=1, value=label)
        cb = ws_sum.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            ca.font = title_font
            cb.font = title_font
        else:
            ca.font = label_font
            cb.font = value_font
            cb.alignment = Alignment(horizontal="left")

    wb.save(output_path)
    log.info("XLSX exported → {p} ({n} rows)", p=output_path, n=len(rows))
    return output_path


def run_export(
    score_min: int = 0,
    zones: Optional[list[str]] = None,
    include_agencies: bool = False,
    exclude_room_rentals: bool = True,
    mobile_only: bool = False,
    fmt: str = "both",        # "csv" | "xlsx" | "both"
    output_dir: Optional[str] = None,
    limit: int = 2000,
) -> dict[str, str]:
    """
    Full export flow: build list → export to CSV and/or XLSX.

    Returns dict with keys 'csv' and/or 'xlsx' pointing to output paths.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = Path(output_dir) if output_dir else settings.data_dir
    base_dir.mkdir(parents=True, exist_ok=True)

    rows = generate_contact_list(
        score_min=score_min,
        zones=zones,
        include_agencies=include_agencies,
        exclude_room_rentals=exclude_room_rentals,
        mobile_only=mobile_only,
        limit=limit,
    )

    if not rows:
        log.warning("No leads matched the export criteria — nothing to export")
        return {}

    results: dict[str, str] = {}

    if fmt in ("csv", "both"):
        csv_path = str(base_dir / f"contactos_{ts}.csv")
        results["csv"] = export_contact_csv(rows, csv_path)

    if fmt in ("xlsx", "both"):
        xlsx_path = str(base_dir / f"contactos_{ts}.xlsx")
        try:
            results["xlsx"] = export_contact_xlsx(rows, xlsx_path)
        except ImportError as e:
            log.warning("XLSX skipped: {e}", e=e)

    return results
