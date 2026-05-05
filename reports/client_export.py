"""
Export no formato exato do cliente — "Directriz scraping.xlsx".

O cliente entregou um briefing com 4 sheets:

  1. Básico       — colunas mínimas de qualquer lead
  2. FSBO_FRBO    — proprietários directos (Idealista venda + OLX arrendamento)
                    em Almada/Seixal/Sesimbra + Lisboa/Oeiras/Cascais/Sintra
                    Categorias: casas, garagens, arrecadações, escritórios,
                    espaços comerciais, armazéns, terrenos, prédios.
  3. Linkedin     — pessoas (PT + EUA + Inglaterra) com email + aniversário
  4. Standvirtual — qualquer venda particular >= 15.000€

Critério comum, definido no sheet "Objectivo":
  > Contacto telefónico de PARTICULARES (pessoas) que tenham colocado o
    imóvel à venda/arrendamento, ou pessoas para qualificação de
    vontade/necessidade de venda.

Este módulo gera UM ficheiro xlsx multi-sheet com este layout exato,
prontos para entregar ao cliente sem edição manual.

Usage:
    from reports.client_export import build_client_xlsx
    path = build_client_xlsx()                  # default location
    path = build_client_xlsx("custom.xlsx")
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from sqlalchemy import desc, or_, select

from storage.database import get_db
from storage.models import Lead
from utils.logger import get_logger

log = get_logger(__name__)


# ── Client zone allowlist ────────────────────────────────────────────────────
# The 7 concelhos the client cares about, plus all Lisbon freguesias which
# resolve to "Lisboa" for the export.

CLIENT_CONCELHOS_PT: frozenset[str] = frozenset([
    "Almada", "Seixal", "Sesimbra",
    "Lisboa", "Oeiras", "Cascais", "Sintra",
])

# Linkedin extras — international targets the client added on top of the
# PT concelhos.
CLIENT_LINKEDIN_GEOS: frozenset[str] = frozenset([
    "Portugal", "EUA", "Estados Unidos", "USA",
    "Inglaterra", "England", "Reino Unido", "UK", "United Kingdom",
])


# ── Internal helpers ─────────────────────────────────────────────────────────

def _norm_zone(s: Optional[str]) -> str:
    """Lowercase + strip + drop common prefixes for a stable comparison."""
    return (s or "").strip().lower()


def _client_zone_match(zone: Optional[str]) -> bool:
    """True if the lead's zone resolves to one of the 7 client concelhos."""
    if not zone:
        return False
    nz = _norm_zone(zone)
    if nz.startswith("lisboa"):
        return True   # captures freguesia drill-downs ("Lisboa-Alvalade", …)
    for target in CLIENT_CONCELHOS_PT:
        if nz == _norm_zone(target):
            return True
    return False


def _first_url(lead: Lead) -> str:
    try:
        for s in lead.sources or []:
            if s.get("url"):
                return s["url"]
    except Exception:
        pass
    return ""


def _is_individual(lead: Lead) -> bool:
    """
    Particular (FSBO/FRBO) check — must have a phone AND not be flagged as
    agency. Bank/auction/promoter listings are excluded.
    """
    if not lead.contact_phone:
        return False
    ot = (lead.owner_type or "").lower()
    if ot in ("agency", "bank", "developer", "auction"):
        return False
    if lead.seller_super_flag:
        return False           # camouflaged agency
    return True


def _imovel_label(lead: Lead) -> str:
    """Human-readable property type label combining typology + property_type."""
    bits: list[str] = []
    if lead.typology:
        bits.append(lead.typology)
    if lead.property_type:
        bits.append(lead.property_type)
    if lead.title and not bits:
        # Fall back to first 60 chars of title when structured fields empty
        return lead.title[:60]
    return " · ".join(bits) or (lead.title or "")[:60]


def _format_eur(v: Optional[float]) -> str:
    if not v:
        return ""
    return f"{int(v):,}".replace(",", " ") + " €"


def _format_date(d: Optional[datetime]) -> str:
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


# ── Row builders (one per sheet) ─────────────────────────────────────────────

# Sheet 1 — Básico: every actionable lead (any source, particular, with phone)
def _basico_rows(leads: Iterable[Lead]) -> list[dict]:
    out: list[dict] = []
    for lead in leads:
        if not _is_individual(lead):
            continue
        if not _client_zone_match(lead.zone):
            continue
        out.append({
            "Data de entrada": _format_date(lead.first_seen_at),
            "Origem (fonte)":  (lead.discovery_source or "").upper(),
            "Nome":            lead.contact_name or "",
            "Contacto":        lead.contact_phone or "",
            "Link":            _first_url(lead),
        })
    return out


# Sheet 2 — FSBO/FRBO: Idealista venda + OLX arrendamento, particulares só
def _fsbo_frbo_rows(leads: Iterable[Lead]) -> list[dict]:
    out: list[dict] = []
    for lead in leads:
        if not _is_individual(lead):
            continue
        if not _client_zone_match(lead.zone):
            continue
        src = (lead.discovery_source or "").lower()
        # Source filter: Idealista (FSBO) or OLX (FRBO) — covers the client's
        # explicit ask. Custojusto/Sapo/Imovirtual excluded from this sheet
        # to keep the deliverable tight.
        if src not in ("idealista", "olx"):
            continue
        # Concelho — strip Lisbon freguesia drill-downs back to "Lisboa"
        concelho = lead.zone or ""
        if concelho.startswith("Lisboa-"):
            concelho = "Lisboa"
        out.append({
            "Data de entrada":   _format_date(lead.first_seen_at),
            "Origem (fonte)":    src.upper(),
            "Nome Proprietário": lead.contact_name or "",
            "Contacto":          lead.contact_phone or "",
            "Imóvel":            _imovel_label(lead),
            "Tipologia":         lead.typology or "",
            "Freguesia":         lead.parish or "",
            "Concelho":          concelho,
            "Valor imóvel":      _format_eur(lead.price),
            "Link":              _first_url(lead),
        })
    return out


# Sheet 3 — LinkedIn: PT concelhos + EUA + Inglaterra, with email + birthday
def _linkedin_rows(leads: Iterable[Lead]) -> list[dict]:
    out: list[dict] = []
    for lead in leads:
        if (lead.discovery_source or "").lower() != "linkedin":
            continue
        if not lead.contact_phone and not lead.contact_email:
            continue
        # Zone filter: PT concelhos OR international targets stored in zone
        zone_norm = _norm_zone(lead.zone)
        country_match = any(
            g.lower() in zone_norm for g in CLIENT_LINKEDIN_GEOS
        )
        if not (_client_zone_match(lead.zone) or country_match):
            continue
        out.append({
            "Data de entrada": _format_date(lead.first_seen_at),
            "Origem (fonte)":  "LINKEDIN",
            "Nome":            lead.contact_name or "",
            "Contacto":        lead.contact_phone or "",
            "Email":           lead.contact_email or "",
            "Aniversário":     lead.birthday or "",
            "Link":            _first_url(lead),
        })
    return out


# Sheet 4 — Standvirtual: any sale ≥ 15 000 €
def _standvirtual_rows(leads: Iterable[Lead], min_price: float = 15_000.0) -> list[dict]:
    out: list[dict] = []
    for lead in leads:
        src = (lead.discovery_source or "").lower()
        if src != "standvirtual":
            continue
        if not lead.contact_phone:
            continue
        # The Standvirtual scraper stores the asset price in product_value
        # (when the listing is a non-real-estate item) or price (when it's
        # a vehicle priced in the same field). Use whichever is set.
        value = lead.product_value or lead.price or 0
        if value < min_price:
            continue
        out.append({
            "Data de entrada":          _format_date(lead.first_seen_at),
            "Origem (fonte)":           "STANDVIRTUAL",
            "Nome":                     lead.contact_name or "",
            "Contacto":                 lead.contact_phone or "",
            "Tipo de produto à venda":  lead.product_title or _imovel_label(lead),
            "Valor do produto":         _format_eur(value),
            "Link":                     _first_url(lead),
        })
    return out


# ── Public entry point ───────────────────────────────────────────────────────

def build_client_xlsx(output_path: Optional[str | Path] = None) -> Path:
    """
    Build the multi-sheet xlsx in the exact format the client requested.

    Returns the absolute Path to the produced file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl não está instalado — `pip install openpyxl`"
        )

    # Load every active lead in one query
    with get_db() as db:
        leads = db.execute(
            select(Lead)
            .where(Lead.archived == False)        # noqa: E712
            .where(Lead.is_demo == False)         # noqa: E712
            .order_by(desc(Lead.score), desc(Lead.first_seen_at))
        ).scalars().all()

    log.info(
        "[client_export] {n} active leads — building 4 sheets",
        n=len(leads),
    )

    sheets: dict[str, list[dict]] = {
        "Básico":       _basico_rows(leads),
        "FSBO_FRBO":    _fsbo_frbo_rows(leads),
        "Linkedin":     _linkedin_rows(leads),
        "Standvirtual": _standvirtual_rows(leads),
    }

    # Resolve output path
    if output_path:
        out = Path(output_path)
    else:
        from config.settings import settings
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        out = settings.data_dir / f"directriz_cliente_{ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)        # drop the default empty sheet

    HEADER_FILL = PatternFill("solid", fgColor="0B1020")
    HEADER_FONT = Font(bold=True, color="34D399", size=11, name="Inter")
    BODY_FONT   = Font(size=10, name="Inter")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        if not rows:
            ws.append(["(sem dados para este filtro)"])
            ws["A1"].font = Font(italic=True, color="64748B", size=10)
            continue

        # Headers
        headers = list(rows[0].keys())
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center")

        # Body
        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        # Column widths — auto-fit roughly
        for col_idx, header in enumerate(headers, start=1):
            cell_lengths = [len(str(header))] + [
                len(str(r.get(header, ""))) for r in rows
            ]
            max_len = max(cell_lengths) if cell_lengths else 12
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # Apply body font to data rows
        for row_idx in range(2, len(rows) + 2):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = BODY_FONT

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(out)

    log.info(
        "[client_export] saved → {p} (Básico={b}, FSBO_FRBO={f}, "
        "LinkedIn={l}, Standvirtual={s})",
        p=out, b=len(sheets["Básico"]), f=len(sheets["FSBO_FRBO"]),
        l=len(sheets["Linkedin"]), s=len(sheets["Standvirtual"]),
    )

    return out
