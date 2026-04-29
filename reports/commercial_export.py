"""
Commercial Export — lista comercial pronta para entrega ao cliente.

Gera um único XLSX com três folhas:
  1. "Lista Premium"    — proprietários directos, mobile/relay, score alto
  2. "Lista Expandida"  — oportunidades adicionais, mais abrangente
  3. "Resumo Executivo" — KPIs prontos para apresentação ao cliente

Critérios Lista Premium (ordenada: mobile primeiro → relay → score desc):
  • telefone válido obrigatório (mobile ou relay preferido; sem fixos 21x/22x)
  • agências excluídas (owner_type="agency" ou lead_type="agency_listing")
  • tipos permitidos: fsbo, frbo, active_owner, unknown
  • score ≥ warm_threshold (configurável, default: settings.warm_score_threshold)
  • zonas-alvo do .env (configurável)
  • deduplicação por telefone (mantém score mais alto)
  • limite configurável (default: 50)

Critérios Lista Expandida:
  • telefone válido obrigatório
  • agências confirmadas excluídas
  • score ≥ (warm_threshold - 10), floor=25
  • exclui fixos Lisboa/Porto (21x/22x) → provável agência
  • remove telefones já presentes na Lista Premium
  • deduplicação por telefone
  • limite configurável (default: 150)

Uso:
  python main.py export-commercial
  python main.py export-commercial --premium-limit 30 --expanded-limit 100
  python main.py export-commercial --zones Lisboa,Cascais
  python main.py export-commercial --output-dir exports/
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

from config.settings import settings
from storage.database import get_db
from storage.models import Lead
from utils.logger import get_logger

# Re-use helpers from contact_export to avoid duplication
from reports.contact_export import (
    _canonical_phone,
    _whatsapp_link,
    _get_url,
    _build_insight,
    _is_likely_agency,
    _format_price,
    _confidence_label,
    _LEAD_TYPE_LABELS,
    _SOURCE_LABELS,
)

log = get_logger(__name__)

# ── Phone type sort priority (mobile first) ───────────────────────────────────
_PHONE_PRIORITY = {"mobile": 0, "relay": 1, "unknown": 2, "landline": 3}

# ── Phone type labels ─────────────────────────────────────────────────────────
_PHONE_LABELS = {
    "mobile":   "Telemóvel",
    "relay":    "Relay/OLX",
    "landline": "Fixo",
    "unknown":  "Desconhecido",
}

# ── Commercial insight — richer than generic insight ─────────────────────────
_URGENCY_RE: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\burgente\b|\burgência\b", re.I),              "venda urgente"),
    (re.compile(r"\bherança\b|\bherdeiro\b|\bpartilha\b", re.I), "herança/partilha"),
    (re.compile(r"\bdivórcio\b|\bseparação\b", re.I),            "divórcio"),
    (re.compile(r"\bemigr\w+\b", re.I),                          "proprietário a emigrar"),
    (re.compile(r"\bpenhora\b|\bexecutiv\w+\b", re.I),           "execução hipotecária"),
    (re.compile(r"\bpermuta\b", re.I),                           "aberto a permuta"),
    (re.compile(r"\bnegocia\w*\b", re.I),                        "preço negociável"),
    (re.compile(r"sem comissão|direto do dono|proprietário vende", re.I),
                                                                  "direto do proprietário"),
    (re.compile(r"\bpreciso vender\b|\bprecisa vender\b", re.I), "precisa vender"),
    (re.compile(r"\bpara recuperar\b|\bpara remodelar\b", re.I), "imóvel para remodelar"),
]


def _commercial_insight(lead: Lead) -> str:
    """
    Build a single-line commercial insight for the client list.

    Format: [tipo_proprietário] · [canal] · [sinal principal]

    Examples:
      "Proprietário Venda · Telemóvel · 23% abaixo do mercado"
      "Proprietário Arrendamento · Telemóvel · senhorio directo"
      "Potencial Proprietário · Relay/OLX · 97 dias no mercado"
    """
    parts: list[str] = []

    # 1. Lead type label
    lt = lead.lead_type or "unknown"
    lbl = _LEAD_TYPE_LABELS.get(lt, lt)
    parts.append(lbl)

    # 2. Phone channel
    pt = getattr(lead, "phone_type", None) or "unknown"
    parts.append(_PHONE_LABELS.get(pt, "Desconhecido"))

    # 3. Primary signal (first match wins)
    signal = ""

    # Price delta takes priority
    delta = lead.price_delta_pct
    if delta and delta >= 15:
        signal = f"{delta:.0f}% abaixo do mercado"
    elif delta and delta >= 5:
        signal = f"{delta:.0f}% abaixo do benchmark"

    # Urgency keywords
    if not signal:
        text = f"{lead.title or ''} {lead.description or ''}"
        for pat, label in _URGENCY_RE:
            if pat.search(text):
                signal = label
                break

    # Days on market
    if not signal:
        dom = lead.days_on_market or 0
        if dom >= 90:
            signal = f"{dom} dias no mercado"
        elif dom >= 45:
            signal = f"{dom} dias listado"

    # FRBO specific
    if not signal and lt == "frbo":
        signal = "senhorio a arrendar directamente"

    # Fallback — mobile confirmation
    if not signal and pt == "mobile":
        signal = "telemóvel directo confirmado"

    if signal:
        parts.append(signal)

    return " · ".join(parts)


def _is_excluded_landline(lead: Lead) -> bool:
    """Return True for Lisboa/Porto landlines — almost always agency lines."""
    phone = (lead.contact_phone or "").replace("+351", "").strip()
    return phone[:2] in ("21", "22")


# ── List builders ─────────────────────────────────────────────────────────────

def _lead_to_row(lead: Lead, rank: int | None = None) -> dict:
    """Convert a Lead ORM object to an export row dict."""
    phone     = _canonical_phone(lead.contact_phone or "") or lead.contact_phone or ""
    pt        = getattr(lead, "phone_type", None) or "unknown"
    tipo_tel  = _PHONE_LABELS.get(pt, "Desconhecido")
    url       = _get_url(lead.sources_json or "[]") or (lead.url if hasattr(lead, "url") else "")
    conf      = _confidence_label(lead)
    insight   = _commercial_insight(lead)
    wa_link   = _whatsapp_link(phone) if phone else ""

    row = {
        "rank":          rank,
        "score":         lead.score or 0,
        "label":         lead.score_label or "COLD",
        "nome":          lead.contact_name or "—",
        "telefone":      phone,
        "tipo_telefone": tipo_tel,
        "whatsapp":      wa_link,
        "zona":          lead.zone or "—",
        "concelho":      lead.municipality or lead.zone or "—",
        "tipologia":     lead.typology or "—",
        "preco":         _format_price(lead.price),
        "area_m2":       f"{lead.area_m2:.0f} m²" if lead.area_m2 else "—",
        "tipo_lead":     _LEAD_TYPE_LABELS.get(lead.lead_type or "unknown", lead.lead_type or "—"),
        "fonte":         _SOURCE_LABELS.get(lead.discovery_source or "", lead.discovery_source or "—"),
        "confianca":     conf,
        "insight":       insight,
        "url":           url,
        "dias_mercado":  lead.days_on_market or 0,
        "data_captacao": lead.first_seen_at.strftime("%d/%m/%Y") if lead.first_seen_at else "—",
        # internal fields for filtering/sorting — not written to sheet
        "_phone_type":   pt,
        "_lead_id":      lead.id,
        "_owner_type":   lead.owner_type or "",
    }
    return row


def generate_premium_list(
    score_min: int | None = None,
    zones: list[str] | None = None,
    limit: int = 50,
    min_confidence_score: int = 0,
) -> list[dict]:
    """
    Build the Lista Premium.

    Strict criteria:
      • score ≥ max(score_min, warm_threshold)
      • owner confirmed or unknown — agencies excluded
      • phone required, valid, not a Lisboa/Porto agency landline
      • FSBO/FRBO/active_owner/unknown lead types
      • target zones only (from .env unless overridden)
      • sorted: mobile first → relay → score desc
      • dedup by phone

    Returns list of row dicts, up to `limit` entries.
    """
    from sqlalchemy import select, and_, or_

    warm_t    = settings.warm_score_threshold
    min_score = max(score_min if score_min is not None else warm_t, warm_t)
    target_zones = zones or settings.zones

    with get_db() as db:
        q = (
            select(Lead)
            .where(
                Lead.archived == False,                       # noqa: E712
                Lead.is_demo  == False,                       # noqa: E712
                Lead.contact_phone.isnot(None),
                Lead.contact_phone != "",
                Lead.score >= min_score,
            )
            .where(
                or_(
                    Lead.owner_type.in_(("fsbo", "unknown", "developer")),
                    Lead.owner_type.is_(None),
                )
            )
            .where(
                Lead.lead_type.notin_(("agency_listing",))
                if True else True
            )
            .where(Lead.zone.in_(target_zones))
            .order_by(Lead.score.desc())
            .limit(limit * 5)   # buffer for dedup + landline filter
        )
        leads = db.execute(q).scalars().all()

    seen_phones: set[str] = set()
    rows: list[dict] = []

    # Sort: adjusted score = real score + phone channel bonus (5 for mobile, 2 for relay).
    # This keeps HOT relay leads above low-score mobile leads, while mobile wins
    # at equal or near-equal scores — matching the user's intent of "mobile first"
    # without demoting high-value relay leads.
    def _sort_key(lead: Lead):
        pt    = getattr(lead, "phone_type", None) or "unknown"
        bonus = 5 if pt == "mobile" else (2 if pt == "relay" else 0)
        return -(( lead.score or 0) + bonus)

    leads_sorted = sorted(leads, key=_sort_key)

    for lead in leads_sorted:
        phone = _canonical_phone(lead.contact_phone or "")
        if not phone:
            continue
        if phone in seen_phones:
            continue
        if _is_excluded_landline(lead):
            continue
        if _is_likely_agency(lead):
            continue
        # Filter active_owner + landline — weak signal
        if lead.lead_type == "active_owner" and (
            getattr(lead, "phone_type", None) == "landline"
        ):
            continue

        seen_phones.add(phone)
        row = _lead_to_row(lead, rank=len(rows) + 1)
        rows.append(row)

        if len(rows) >= limit:
            break

    log.info(
        "[commercial] Premium list: {n} leads (from {total} candidates)",
        n=len(rows), total=len(leads),
    )
    return rows


def generate_expanded_list(
    premium_phones: set[str],
    score_min: int | None = None,
    zones: list[str] | None = None,
    limit: int = 150,
) -> list[dict]:
    """
    Build the Lista Expandida.

    Broader criteria than Premium:
      • score ≥ max(score_min, warm_threshold - 10, 25)
      • agencies excluded
      • Lisboa/Porto landlines excluded
      • phones already in Premium excluded
      • dedup by phone
      • sorted by score desc

    Returns list of row dicts, up to `limit` entries.
    """
    from sqlalchemy import select, or_

    warm_t    = settings.warm_score_threshold
    min_score = max(score_min if score_min is not None else warm_t - 10, 25)
    target_zones = zones or settings.zones

    with get_db() as db:
        q = (
            select(Lead)
            .where(
                Lead.archived == False,                       # noqa: E712
                Lead.is_demo  == False,                       # noqa: E712
                Lead.contact_phone.isnot(None),
                Lead.contact_phone != "",
                Lead.score >= min_score,
            )
            .where(
                or_(
                    Lead.owner_type != "agency",
                    Lead.owner_type.is_(None),
                )
            )
            .where(Lead.lead_type.notin_(("agency_listing",)))
            .where(Lead.zone.in_(target_zones))
            .order_by(Lead.score.desc())
            .limit(limit * 4)
        )
        leads = db.execute(q).scalars().all()

    seen_phones: set[str] = set(premium_phones)   # start from premium exclusions
    rows: list[dict] = []

    for lead in leads:
        phone = _canonical_phone(lead.contact_phone or "")
        if not phone:
            continue
        if phone in seen_phones:
            continue
        if _is_excluded_landline(lead):
            continue
        if _is_likely_agency(lead):
            continue

        seen_phones.add(phone)
        row = _lead_to_row(lead, rank=len(rows) + 1)
        rows.append(row)

        if len(rows) >= limit:
            break

    log.info(
        "[commercial] Expanded list: {n} leads (from {total} candidates)",
        n=len(rows), total=len(leads),
    )
    return rows


# ── Executive summary ─────────────────────────────────────────────────────────

def build_executive_summary(
    premium: list[dict],
    expanded: list[dict],
    generated_at: datetime | None = None,
) -> dict:
    """
    Build a structured executive summary from the two lists.

    Returns a dict suitable for both text rendering and XLSX Sheet 3.
    """
    generated_at = generated_at or datetime.now()
    all_rows = premium + expanded

    # Counts
    p_hot     = sum(1 for r in premium  if r["label"] == "HOT")
    p_warm    = sum(1 for r in premium  if r["label"] == "WARM")
    e_hot     = sum(1 for r in expanded if r["label"] == "HOT")
    e_warm    = sum(1 for r in expanded if r["label"] == "WARM")

    # Phone type breakdown (combined)
    pt_counts: Counter = Counter(r["_phone_type"] for r in all_rows)

    # Lead type breakdown
    lt_counts: Counter = Counter(r["tipo_lead"] for r in all_rows)

    # Top zones (premium only — that's the target)
    zone_counts: Counter = Counter(r["zona"] for r in premium)
    top_zones = zone_counts.most_common(6)

    # Score range in premium
    if premium:
        p_scores = [r["score"] for r in premium]
        p_score_min, p_score_max, p_score_avg = (
            min(p_scores),
            max(p_scores),
            round(sum(p_scores) / len(p_scores), 1),
        )
    else:
        p_score_min = p_score_max = p_score_avg = 0

    return {
        "generated_at":      generated_at.strftime("%d/%m/%Y %H:%M"),
        "total_premium":     len(premium),
        "total_expanded":    len(expanded),
        "total_combined":    len(all_rows),
        # Premium breakdown
        "premium_hot":       p_hot,
        "premium_warm":      p_warm,
        "premium_cold":      len(premium) - p_hot - p_warm,
        "premium_score_min": p_score_min,
        "premium_score_max": p_score_max,
        "premium_score_avg": p_score_avg,
        # Expanded breakdown
        "expanded_hot":      e_hot,
        "expanded_warm":     e_warm,
        "expanded_cold":     len(expanded) - e_hot - e_warm,
        # Phone channels
        "mobile_count":      pt_counts.get("mobile", 0),
        "relay_count":       pt_counts.get("relay", 0),
        "landline_count":    pt_counts.get("landline", 0),
        "unknown_count":     pt_counts.get("unknown", 0),
        # Lead types
        "lead_type_counts":  dict(lt_counts.most_common()),
        # Zones
        "top_zones":         top_zones,
    }


def summary_as_text(summary: dict) -> str:
    """Render the executive summary as a human-readable text block."""
    sep  = "─" * 58
    sep2 = "═" * 58
    top_zones_str = "  ".join(
        f"{z} ({n})" for z, n in summary["top_zones"]
    )

    lt = summary["lead_type_counts"]
    lt_str = "  ".join(f"{k}: {v}" for k, v in lt.items())

    lines = [
        "",
        sep2,
        "  RELATÓRIO COMERCIAL — LISTA DE LEADS",
        f"  Gerado em: {summary['generated_at']}",
        sep2,
        "",
        f"  LISTA PREMIUM ({summary['total_premium']} contactos únicos)",
        sep,
        f"  🔴 HOT          : {summary['premium_hot']}",
        f"  🟡 WARM         : {summary['premium_warm']}",
        f"  Score (min/avg/max) : {summary['premium_score_min']} / "
        f"{summary['premium_score_avg']} / {summary['premium_score_max']}",
        "",
        f"  LISTA EXPANDIDA ({summary['total_expanded']} contactos únicos)",
        sep,
        f"  🔴 HOT          : {summary['expanded_hot']}",
        f"  🟡 WARM         : {summary['expanded_warm']}",
        "",
        "  CANAIS DE CONTACTO (combinado)",
        sep,
        f"  📱 Telemóvel    : {summary['mobile_count']}",
        f"  🔁 Relay/OLX    : {summary['relay_count']}",
        f"  📞 Fixo         : {summary['landline_count']}",
        f"  ❓ Desconhecido : {summary['unknown_count']}",
        "",
        "  TIPOS DE LEAD",
        sep,
        f"  {lt_str}",
        "",
        "  PRINCIPAIS ZONAS (Premium)",
        sep,
        f"  {top_zones_str}",
        "",
        sep2,
    ]
    return "\n".join(lines)


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_commercial_xlsx(
    premium:  list[dict],
    expanded: list[dict],
    summary:  dict,
    output_path: str,
) -> str:
    """
    Generate a single client-ready XLSX with three sheets.

    Sheet 1: Lista Premium
    Sheet 2: Lista Expandida
    Sheet 3: Resumo Executivo
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────────────────
    C = {
        "header_bg":    "1F3864",
        "header_fg":    "FFFFFF",
        "hot_bg":       "FFD6D6",
        "warm_bg":      "FFF4CC",
        "mobile_bg":    "E8F5E9",   # soft green — mobile numbers
        "relay_bg":     "E3F2FD",   # soft blue — relay numbers
        "alt_bg":       "F8FBFF",
        "white":        "FFFFFF",
        "border":       "D0D7E0",
        "premium_tab":  "C62828",
        "expanded_tab": "1565C0",
        "summary_tab":  "1B5E20",
    }

    thin = Border(
        left=Side(style="thin",  color=C["border"]),
        right=Side(style="thin", color=C["border"]),
        top=Side(style="thin",   color=C["border"]),
        bottom=Side(style="thin",color=C["border"]),
    )

    # ── Column definitions (shared between Premium and Expandida) ─────────────
    COLUMNS = [
        ("#",            "rank",          4),
        ("Score",        "score",         7),
        ("Label",        "label",         8),
        ("Nome",         "nome",         22),
        ("Telefone",     "telefone",     16),
        ("Tipo Tel.",    "tipo_telefone",12),
        ("WhatsApp",     "whatsapp",     18),
        ("Zona",         "zona",         12),
        ("Tipologia",    "tipologia",    10),
        ("Preço",        "preco",        14),
        ("Área",         "area_m2",       9),
        ("Tipo Lead",    "tipo_lead",    22),
        ("Fonte",        "fonte",        14),
        ("Confiança",    "confianca",    10),
        ("Insight",      "insight",      42),
        ("URL Anúncio",  "url",          18),
        ("Dias Merc.",   "dias_mercado",  9),
        ("Captado em",   "data_captacao",12),
    ]

    def _write_sheet(ws, rows: list[dict], title: str, tab_color: str) -> None:
        ws.title = title
        ws.sheet_properties.tabColor = tab_color

        # Header
        h_font  = Font(name="Calibri", bold=True, color=C["header_fg"], size=10)
        h_fill  = PatternFill("solid", fgColor=C["header_bg"])
        h_align = Alignment(horizontal="center", vertical="center")

        for ci, (hdr, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font  = h_font
            cell.fill  = h_fill
            cell.border = thin
            cell.alignment = h_align
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Data rows
        d_font    = Font(name="Calibri", size=9)
        bold_font = Font(name="Calibri", size=9, bold=True)
        link_font = Font(name="Calibri", size=9, color="1565C0", underline="single")
        c_align   = Alignment(horizontal="center", vertical="center")
        l_align   = Alignment(horizontal="left",   vertical="center")
        r_align   = Alignment(horizontal="right",  vertical="center")
        wrap_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for ri, row in enumerate(rows, 2):
            pt    = row.get("_phone_type", "unknown")
            label = row.get("label", "COLD")

            # Row background: HOT > WARM > mobile > relay > alternating
            if label == "HOT":
                bg = C["hot_bg"]
            elif label == "WARM":
                bg = C["warm_bg"]
            elif pt == "mobile":
                bg = C["mobile_bg"]
            elif pt == "relay":
                bg = C["relay_bg"]
            elif ri % 2 == 0:
                bg = C["alt_bg"]
            else:
                bg = C["white"]

            row_fill = PatternFill("solid", fgColor=bg)

            for ci, (_, key, _) in enumerate(COLUMNS, 1):
                v    = row.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
                cell.fill   = row_fill

                if key == "rank":
                    cell.font = Font(name="Calibri", size=9, color="888888")
                    cell.alignment = c_align
                elif key == "score":
                    cell.font      = bold_font
                    cell.alignment = r_align
                elif key == "label":
                    color = {"HOT": "C00000", "WARM": "B8860B"}.get(label, "444444")
                    cell.font      = Font(name="Calibri", size=9, bold=True, color=color)
                    cell.alignment = c_align
                elif key == "tipo_telefone":
                    icon = {"Telemóvel": "📱", "Relay/OLX": "🔁",
                            "Fixo": "📞", "Desconhecido": "❓"}.get(str(v), "")
                    cell.value     = f"{icon} {v}" if icon else v
                    cell.font      = d_font
                    cell.alignment = c_align
                elif key == "whatsapp":
                    if v and v.startswith("http"):
                        cell.hyperlink = v
                        cell.value     = "📲 WhatsApp"
                        cell.font      = link_font
                    cell.alignment = c_align
                elif key == "url":
                    if v and v.startswith("http"):
                        cell.hyperlink = v
                        cell.value     = "🔗 Ver Anúncio"
                        cell.font      = link_font
                    cell.alignment = c_align
                elif key == "insight":
                    cell.font      = Font(name="Calibri", size=9, italic=True)
                    cell.alignment = wrap_al
                elif key == "confianca":
                    color = {"ALTA": "1B5E20", "MÉDIA": "E65100", "BAIXA": "B71C1C"}.get(str(v), "444444")
                    cell.font      = Font(name="Calibri", size=9, bold=True, color=color)
                    cell.alignment = c_align
                elif key == "dias_mercado":
                    cell.font      = d_font
                    cell.alignment = r_align
                else:
                    cell.font      = d_font
                    cell.alignment = l_align

            ws.row_dimensions[ri].height = 18

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 1: Premium ──────────────────────────────────────────────────────
    ws_premium = wb.active
    _write_sheet(ws_premium, premium, "Lista Premium", C["premium_tab"])

    # ── Sheet 2: Expandida ────────────────────────────────────────────────────
    ws_expanded = wb.create_sheet()
    _write_sheet(ws_expanded, expanded, "Lista Expandida", C["expanded_tab"])

    # ── Sheet 3: Resumo Executivo ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumo Executivo")
    ws_sum.sheet_properties.tabColor = C["summary_tab"]

    title_font  = Font(name="Calibri", bold=True, size=14, color=C["header_bg"])
    section_font= Font(name="Calibri", bold=True, size=10, color="1F3864")
    label_font  = Font(name="Calibri", size=10)
    value_font  = Font(name="Calibri", bold=True, size=10)
    note_font   = Font(name="Calibri", size=9, italic=True, color="666666")

    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 28

    def _sum_row(r, col_a, col_b="", col_c="", bold=False, section=False, note=False):
        ca = ws_sum.cell(row=r, column=1, value=col_a)
        cb = ws_sum.cell(row=r, column=2, value=col_b)
        cc = ws_sum.cell(row=r, column=3, value=col_c)
        if section:
            ca.font = section_font
            ca.fill = PatternFill("solid", fgColor="EEF2F8")
        elif bold:
            ca.font = value_font
            cb.font = value_font
        elif note:
            ca.font = note_font
            cb.font = note_font
        else:
            ca.font = label_font
            cb.font = value_font
            cc.font = label_font
        return r + 1

    r = 1
    ws_sum.cell(row=r, column=1, value="Relatório Comercial de Leads").font = title_font
    r += 1
    ws_sum.cell(row=r, column=1, value=f"Gerado em: {summary['generated_at']}").font = note_font
    r += 2

    r = _sum_row(r, "LISTA PREMIUM", section=True)
    r = _sum_row(r, "Total contactos únicos",   summary["total_premium"])
    r = _sum_row(r, "  🔴 HOT (score ≥ 60)",   summary["premium_hot"])
    r = _sum_row(r, "  🟡 WARM (score 40-59)",  summary["premium_warm"])
    r = _sum_row(r, "  Score min / avg / max",
                 f"{summary['premium_score_min']} / {summary['premium_score_avg']} / {summary['premium_score_max']}")
    r += 1

    r = _sum_row(r, "LISTA EXPANDIDA", section=True)
    r = _sum_row(r, "Total contactos únicos",   summary["total_expanded"])
    r = _sum_row(r, "  🔴 HOT",                 summary["expanded_hot"])
    r = _sum_row(r, "  🟡 WARM",                summary["expanded_warm"])
    r += 1

    r = _sum_row(r, "CANAIS DE CONTACTO (combinado)", section=True)
    r = _sum_row(r, "📱 Telemóvel",    summary["mobile_count"])
    r = _sum_row(r, "🔁 Relay/OLX",   summary["relay_count"])
    r = _sum_row(r, "📞 Fixo",         summary["landline_count"])
    r = _sum_row(r, "❓ Desconhecido", summary["unknown_count"])
    r += 1

    r = _sum_row(r, "TIPOS DE LEAD", section=True)
    for lt_label, lt_count in summary["lead_type_counts"].items():
        r = _sum_row(r, f"  {lt_label}", lt_count)
    r += 1

    r = _sum_row(r, "PRINCIPAIS ZONAS (Premium)", section=True)
    for zone, count in summary["top_zones"]:
        bar = "█" * min(count, 20)
        r = _sum_row(r, f"  {zone}", count, bar)
    r += 1

    r = _sum_row(r, "TOTAL GERAL", summary["total_combined"], bold=True)

    r += 2
    _sum_row(r, "Legenda cores Premium/Expandida:", note=True)
    r += 1
    _sum_row(r, "  Vermelho claro = HOT (score ≥ 60)", note=True)
    r += 1
    _sum_row(r, "  Amarelo claro = WARM (score 40–59)", note=True)
    r += 1
    _sum_row(r, "  Verde claro = Telemóvel directo", note=True)
    r += 1
    _sum_row(r, "  Azul claro = Relay/OLX", note=True)

    wb.save(output_path)
    log.info("[commercial] XLSX saved → {p}", p=output_path)
    return output_path


# ── CSV export (fallback / two separate files) ────────────────────────────────

def export_commercial_csv(
    rows: list[dict],
    output_path: str,
) -> str:
    """Export a single list to CSV (utf-8-sig for Excel compatibility)."""
    import csv
    fields = [
        "rank", "score", "label", "nome", "telefone", "tipo_telefone",
        "whatsapp", "zona", "concelho", "tipologia", "preco", "area_m2",
        "tipo_lead", "fonte", "confianca", "insight", "url",
        "dias_mercado", "data_captacao",
    ]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    log.info("[commercial] CSV saved → {p} ({n} rows)", p=output_path, n=len(rows))
    return output_path


# ── Main orchestrator ─────────────────────────────────────────────────────────

def run_commercial_export(
    premium_limit:  int = 50,
    expanded_limit: int = 150,
    zones: list[str] | None = None,
    fmt: str = "xlsx",           # "xlsx" | "csv" | "both"
    output_dir: str | None = None,
) -> dict:
    """
    Full commercial export flow.

    1. Build Lista Premium
    2. Build Lista Expandida (excluding Premium phones)
    3. Build Executive Summary
    4. Export to XLSX (single file, 3 sheets) and/or CSV (2 files)

    Returns:
        {
            "premium":  list[dict],
            "expanded": list[dict],
            "summary":  dict,
            "files":    {"xlsx": str, "csv_premium": str, "csv_expanded": str},
        }
    """
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(output_dir) if output_dir else Path(settings.data_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    log.info("[commercial] Building Premium list (limit={n})…", n=premium_limit)
    premium = generate_premium_list(zones=zones, limit=premium_limit)

    premium_phones = {r["telefone"] for r in premium if r.get("telefone")}

    log.info("[commercial] Building Expanded list (limit={n})…", n=expanded_limit)
    expanded = generate_expanded_list(
        premium_phones=premium_phones,
        zones=zones,
        limit=expanded_limit,
    )

    summary = build_executive_summary(premium, expanded)

    log.info(
        "[commercial] Summary — Premium: {p} | Expanded: {e} | "
        "Mobile: {m} | Relay: {r}",
        p=len(premium), e=len(expanded),
        m=summary["mobile_count"], r=summary["relay_count"],
    )

    files: dict[str, str] = {}

    if fmt in ("xlsx", "both"):
        xlsx_path = str(out_dir / f"leads_comercial_{ts}.xlsx")
        try:
            export_commercial_xlsx(premium, expanded, summary, xlsx_path)
            files["xlsx"] = xlsx_path
        except ImportError as exc:
            log.warning("XLSX export skipped: {e}", e=exc)

    if fmt in ("csv", "both"):
        csv_p = str(out_dir / f"leads_premium_{ts}.csv")
        csv_e = str(out_dir / f"leads_expandida_{ts}.csv")
        export_commercial_csv(premium,  csv_p)
        export_commercial_csv(expanded, csv_e)
        files["csv_premium"]  = csv_p
        files["csv_expanded"] = csv_e

    return {
        "premium":  premium,
        "expanded": expanded,
        "summary":  summary,
        "files":    files,
    }
