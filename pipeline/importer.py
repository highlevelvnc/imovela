"""
Lead Importer — normalise and import CSV/XLSX files into the LeadEngine database.

Handles both PT-labelled exports (from export-contacts/export-commercial)
and English-column files.  Auto-detects column mapping from headers.

Usage:
    from pipeline.importer import LeadImporter
    stats = LeadImporter().import_file("data/contactos_20260403.csv")
    stats = LeadImporter(dry_run=True).import_directory("data/", "contactos_*.csv")
"""
from __future__ import annotations

import csv
import glob
import json
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from utils.logger import get_logger

log = get_logger(__name__)


# ── Column mapping: header → Lead field ──────────────────────────────────────
# Keys are lowercase+stripped.  Supports both PT (export-contacts) and EN
# (raw export) headers.  First match wins.

_COLUMN_MAP: dict[str, str] = {
    # PT export columns (from export-contacts)
    "nome":           "contact_name",
    "telefone":       "contact_phone",
    "tipo_telefone":  "_tipo_telefone",   # mapped post-hoc to phone_type
    "whatsapp":       "_whatsapp_url",    # extract phone from wa.me URL
    "zona":           "zone",
    "concelho":       "municipality",
    "tipologia":      "typology",
    "preco":          "_price_raw",
    "area_m2":        "_area_raw",
    "tipo_lead":      "_tipo_lead_pt",
    "fonte":          "_fonte_pt",
    "agencia":        "_agencia_pt",
    "confianca":      "_skip",
    "insight":        "_skip",
    "titulo":         "title",
    "url":            "url",
    "data":           "_data_pt",
    "dias_mercado":   "days_on_market",
    # PT export columns (from export-commercial)
    "nome_completo":  "contact_name",
    "primeiro_nome":  "first_name",
    "apelido":        "last_name",
    "email":          "contact_email",
    "aniversario":    "birthday",
    "aniversário":    "birthday",
    "freguesia":      "parish",
    "morada":         "address",
    "tipo_imovel":    "property_type",
    "tipo_imóvel":    "property_type",
    "origem":         "_fonte_pt",
    "valor":          "_price_raw",
    # English columns (from raw export / JSON)
    "contact_name":   "contact_name",
    "contact_phone":  "contact_phone",
    "contact_email":  "contact_email",
    "zone":           "zone",
    "municipality":   "municipality",
    "parish":         "parish",
    "address":        "address",
    "typology":       "typology",
    "price":          "price",
    "area":           "_area_raw",
    "title":          "title",
    "description":    "description",
    "score":          "_score_raw",
    "score_label":    "_skip",
    "label":          "_skip",
    "is_owner":       "is_owner",
    "owner_type":     "owner_type",
    "lead_type":      "lead_type",
    "agency_name":    "agency_name",
    "condition":      "condition",
    "property_type":  "property_type",
    "discovery_source": "discovery_source",
    "first_name":     "first_name",
    "last_name":      "last_name",
    "birthday":       "birthday",
    "phone_type":     "phone_type",
    "days_on_market": "days_on_market",
}

# Lead type PT labels → internal codes
_LEAD_TYPE_MAP: dict[str, str] = {
    "proprietário venda":       "fsbo",
    "proprietario venda":       "fsbo",
    "fsbo":                     "fsbo",
    "senhorio arrendamento":    "frbo",
    "frbo":                     "frbo",
    "active owner":             "active_owner",
    "proprietário activo":      "active_owner",
    "proprietario activo":      "active_owner",
    "agência":                  "agency_listing",
    "agencia":                  "agency_listing",
    "agency":                   "agency_listing",
    "promotor":                 "developer_listing",
    "developer":                "developer_listing",
}

# Source PT labels → internal codes
_SOURCE_MAP: dict[str, str] = {
    "olx":          "olx",
    "imovirtual":   "imovirtual",
    "idealista":    "idealista",
    "standvirtual": "standvirtual",
    "custojusto":   "custojusto",
    "sapo":         "sapo",
    "linkedin":     "linkedin",
    "sapo casa":    "sapo",
}


@dataclass
class ImportStats:
    files_read:  int = 0
    rows_read:   int = 0
    rows_skipped: int = 0
    created:     int = 0
    updated:     int = 0
    errors:      int = 0
    dry_run:     bool = False

    @property
    def total_processed(self) -> int:
        return self.created + self.updated

    def as_text(self) -> str:
        mode = " [DRY RUN]" if self.dry_run else ""
        return (
            f"── Import{mode} ──\n"
            f"  Ficheiros lidos   : {self.files_read}\n"
            f"  Linhas lidas      : {self.rows_read}\n"
            f"  Linhas ignoradas  : {self.rows_skipped}\n"
            f"  Leads criados     : {self.created}\n"
            f"  Leads actualizados: {self.updated}\n"
            f"  Erros             : {self.errors}"
        )


class LeadImporter:
    """Import CSV/XLSX files into the LeadEngine database."""

    def __init__(self, dry_run: bool = False, score_after: bool = True):
        self.dry_run = dry_run
        self.score_after = score_after

    def import_file(self, path: str, sheet: str = None) -> ImportStats:
        """Import a single CSV or XLSX file."""
        stats = ImportStats(dry_run=self.dry_run)
        path = str(path)

        if not os.path.isfile(path):
            log.error("[import] File not found: {p}", p=path)
            return stats

        ext = Path(path).suffix.lower()
        if ext == ".csv":
            rows = self._read_csv(path)
        elif ext in (".xlsx", ".xls"):
            rows = self._read_xlsx(path, sheet)
        else:
            log.error("[import] Unsupported file type: {e}", e=ext)
            return stats

        stats.files_read = 1
        stats.rows_read = len(rows)
        log.info("[import] {p} → {n} rows", p=os.path.basename(path), n=len(rows))

        self._process_rows(rows, stats)

        if self.score_after and not self.dry_run and stats.total_processed > 0:
            self._rescore()

        return stats

    def import_directory(self, dir_path: str, pattern: str = "*.csv") -> ImportStats:
        """Import all matching files in a directory."""
        stats = ImportStats(dry_run=self.dry_run)
        full_pattern = os.path.join(dir_path, pattern)
        files = sorted(glob.glob(full_pattern))

        if not files:
            log.warning("[import] No files matching {p}", p=full_pattern)
            return stats

        log.info("[import] Found {n} files matching {p}", n=len(files), p=pattern)

        for fpath in files:
            file_stats = self.import_file(fpath)
            stats.files_read += file_stats.files_read
            stats.rows_read += file_stats.rows_read
            stats.rows_skipped += file_stats.rows_skipped
            stats.created += file_stats.created
            stats.updated += file_stats.updated
            stats.errors += file_stats.errors

        # Score once at the end
        if self.score_after and not self.dry_run and stats.total_processed > 0:
            self._rescore()

        return stats

    def normalize_and_export(
        self, dir_path: str, pattern: str, output_path: str
    ) -> ImportStats:
        """
        Read all matching files, normalize, deduplicate, and export
        a single clean file (CSV or XLSX based on output extension).
        """
        stats = ImportStats(dry_run=True)
        full_pattern = os.path.join(dir_path, pattern)
        files = sorted(glob.glob(full_pattern))

        if not files:
            log.warning("[normalize-export] No files matching {p}", p=full_pattern)
            return stats

        all_rows: list[dict] = []
        for fpath in files:
            ext = Path(fpath).suffix.lower()
            if ext == ".csv":
                rows = self._read_csv(fpath)
            elif ext in (".xlsx", ".xls"):
                rows = self._read_xlsx(fpath)
            else:
                continue
            stats.files_read += 1
            stats.rows_read += len(rows)
            all_rows.extend(rows)

        # Normalize all rows
        normalized: list[dict] = []
        seen_fps: set[str] = set()
        for row in all_rows:
            try:
                mapped = self._map_row(row)
                if not mapped:
                    stats.rows_skipped += 1
                    continue
                norm = self._normalize_row(mapped)
                fp = self._compute_fingerprint(norm)
                if fp in seen_fps:
                    stats.rows_skipped += 1
                    continue
                seen_fps.add(fp)
                normalized.append(norm)
            except Exception as e:
                stats.errors += 1
                log.debug("[normalize-export] row error: {e}", e=e)

        stats.created = len(normalized)
        log.info(
            "[normalize-export] {n} unique leads from {f} files ({s} skipped/dupes)",
            n=len(normalized), f=stats.files_read, s=stats.rows_skipped,
        )

        # Export
        out_ext = Path(output_path).suffix.lower()
        if out_ext == ".csv":
            self._export_csv(normalized, output_path)
        elif out_ext in (".xlsx", ".xls"):
            self._export_xlsx(normalized, output_path)
        else:
            self._export_csv(normalized, output_path)

        log.info("[normalize-export] Exported to {p}", p=output_path)
        return stats

    # ── File readers ─────────────────────────────────────────────────────────

    @staticmethod
    def _read_csv(path: str) -> list[dict]:
        """Read a CSV file, handling UTF-8 BOM and various delimiters."""
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            for delimiter in (",", ";", "\t"):
                try:
                    with open(path, newline="", encoding=encoding) as f:
                        reader = csv.DictReader(f, delimiter=delimiter)
                        rows = list(reader)
                        # Validate: a good delimiter produces multiple columns
                        if rows and len(rows[0]) > 2:
                            return rows
                except (UnicodeDecodeError, csv.Error):
                    continue
        log.error("[import] Cannot read CSV: {p}", p=path)
        return []

    @staticmethod
    def _read_xlsx(path: str, sheet: str = None) -> list[dict]:
        """Read an XLSX file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            log.error("[import] openpyxl not installed — pip install openpyxl")
            return []

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)

        header = next(rows_iter, None)
        if not header:
            wb.close()
            return []

        header = [str(h or "").strip() for h in header]
        result = []
        for row in rows_iter:
            d = {}
            for i, val in enumerate(row):
                if i < len(header) and header[i]:
                    d[header[i]] = str(val) if val is not None else ""
            if any(v.strip() for v in d.values()):
                result.append(d)
        wb.close()
        return result

    # ── Row mapping & normalisation ──────────────────────────────────────────

    @staticmethod
    def _map_row(row: dict) -> Optional[dict]:
        """Map CSV/XLSX column names to internal Lead field names."""
        mapped: dict[str, str] = {}
        for col_name, value in row.items():
            clean_col = col_name.strip().lower().replace("\ufeff", "")
            target = _COLUMN_MAP.get(clean_col)
            if target and target != "_skip":
                mapped[target] = (value or "").strip()
        return mapped if mapped else None

    @staticmethod
    def _normalize_row(mapped: dict) -> dict:
        """Apply field-level normalisation to a mapped row."""
        from utils.helpers import (
            split_pt_name, parse_price, parse_area,
            normalise_zone, extract_typology,
        )
        from utils.phone import validate_pt_phone

        result: dict = {}

        # Title
        result["title"] = mapped.get("title", "")

        # Contact name → split
        name = mapped.get("contact_name", "")
        result["contact_name"] = name
        if not mapped.get("first_name"):
            first, last = split_pt_name(name)
            result["first_name"] = first or None
            result["last_name"] = last or None
        else:
            result["first_name"] = mapped.get("first_name") or None
            result["last_name"] = mapped.get("last_name") or None

        # Phone
        raw_phone = mapped.get("contact_phone", "")
        if raw_phone:
            pr = validate_pt_phone(raw_phone)
            if pr.valid:
                result["contact_phone"] = pr.canonical
                result["phone_type"] = mapped.get("phone_type") or pr.phone_type
            else:
                result["contact_phone"] = ""
                result["phone_type"] = "unknown"
        else:
            result["contact_phone"] = ""
            result["phone_type"] = "unknown"

        # WhatsApp URL → extract phone
        wa_url = mapped.get("_whatsapp_url", "")
        if wa_url and not result["contact_phone"]:
            m = re.search(r"351(\d{9})", wa_url)
            if m:
                pr = validate_pt_phone(m.group(1))
                if pr.valid:
                    result["contact_phone"] = pr.canonical
                    result["phone_type"] = pr.phone_type

        # Email
        result["contact_email"] = mapped.get("contact_email", "") or None
        result["birthday"] = mapped.get("birthday", "") or None

        # Price
        price_raw = mapped.get("_price_raw") or mapped.get("price", "")
        result["price"] = parse_price(price_raw)

        # Area
        area_raw = mapped.get("_area_raw") or mapped.get("area_m2", "")
        result["area_m2"] = parse_area(area_raw)

        # Zone
        result["zone"] = normalise_zone(mapped.get("zone", ""))
        result["municipality"] = mapped.get("municipality", "") or result["zone"]
        result["parish"] = mapped.get("parish", "") or None
        result["address"] = mapped.get("address", "") or None

        # Typology
        result["typology"] = mapped.get("typology") or extract_typology(
            result["title"]
        )
        result["property_type"] = mapped.get("property_type", "") or None

        # Lead type (PT label → code)
        tipo_lead_pt = (mapped.get("_tipo_lead_pt") or "").lower().strip()
        result["lead_type"] = (
            mapped.get("lead_type")
            or _LEAD_TYPE_MAP.get(tipo_lead_pt, "unknown")
        )

        # Source
        fonte_pt = (mapped.get("_fonte_pt") or "").lower().strip()
        result["discovery_source"] = (
            mapped.get("discovery_source")
            or _SOURCE_MAP.get(fonte_pt, fonte_pt or "import")
        )

        # Owner / agency
        agencia_pt = (mapped.get("_agencia_pt") or "").strip().lower()
        if mapped.get("is_owner") is not None:
            result["is_owner"] = str(mapped["is_owner"]).lower() in ("true", "1", "yes", "sim")
        elif agencia_pt in ("não", "nao", "no", ""):
            result["is_owner"] = True
        else:
            result["is_owner"] = False

        if result["is_owner"]:
            result["owner_type"] = mapped.get("owner_type") or "fsbo"
            result["agency_name"] = None
        else:
            result["owner_type"] = mapped.get("owner_type") or "agency"
            agency_val = mapped.get("_agencia_pt") or mapped.get("agency_name", "")
            result["agency_name"] = agency_val if agency_val.lower() not in ("não", "nao", "no", "") else None

        # Description, condition
        result["description"] = mapped.get("description", "") or result["title"]
        result["condition"] = mapped.get("condition", "") or None

        # URL → sources_json
        url = mapped.get("url", "")
        if url:
            result["url"] = url
            result["sources_json"] = json.dumps([{
                "source": result["discovery_source"],
                "url": url,
                "seen_at": datetime.utcnow().isoformat(),
            }], default=str)

        # Days on market
        dom = mapped.get("days_on_market", "")
        if dom and dom.isdigit():
            result["days_on_market"] = int(dom)

        # Contact source + confidence
        result["contact_source"] = "import"
        if result["contact_phone"]:
            result["contact_confidence"] = 100
        elif result.get("contact_email"):
            result["contact_confidence"] = 70
        elif result["contact_name"]:
            result["contact_confidence"] = 30
        else:
            result["contact_confidence"] = 0

        return result

    @staticmethod
    def _compute_fingerprint(norm: dict) -> str:
        """Compute a dedup fingerprint matching the pipeline's logic."""
        from utils.helpers import fingerprint, slugify_text

        typology = (norm.get("typology") or "").lower()
        zone = (norm.get("zone") or "").lower()
        price = norm.get("price") or 0
        price_rounded = round(price / 1000) * 1000
        area = norm.get("area_m2") or 0
        area_rounded = round(area / 5) * 5
        title = norm.get("title") or ""
        title_slug = slugify_text(" ".join(title.split()[:6]))
        return fingerprint(typology, zone, str(price_rounded), str(area_rounded), title_slug)

    # ── Database operations ──────────────────────────────────────────────────

    def _process_rows(self, rows: list[dict], stats: ImportStats) -> None:
        """Map, normalise, and upsert each row."""
        from storage.database import get_db
        from storage.repository import LeadRepo

        for i, row in enumerate(rows):
            try:
                mapped = self._map_row(row)
                if not mapped:
                    stats.rows_skipped += 1
                    continue

                norm = self._normalize_row(mapped)
                if not norm.get("title") and not norm.get("contact_phone"):
                    stats.rows_skipped += 1
                    continue

                fp = self._compute_fingerprint(norm)

                if self.dry_run:
                    log.debug(
                        "[import] [DRY] #{i} fp={fp} {t:.40s} zone={z} phone={p}",
                        i=i + 1, fp=fp[:8], t=norm.get("title", "")[:40],
                        z=norm.get("zone", "?"), p=norm.get("contact_phone", ""),
                    )
                    stats.created += 1
                    continue

                # Build lead data dict
                lead_data = {
                    "fingerprint":       fp,
                    "is_demo":           False,
                    "title":             norm.get("title"),
                    "typology":          norm.get("typology"),
                    "property_type":     norm.get("property_type"),
                    "area_m2":           norm.get("area_m2"),
                    "price":             norm.get("price"),
                    "zone":              norm.get("zone"),
                    "municipality":      norm.get("municipality"),
                    "parish":            norm.get("parish"),
                    "address":           norm.get("address"),
                    "description":       norm.get("description"),
                    "contact_name":      norm.get("contact_name"),
                    "first_name":        norm.get("first_name"),
                    "last_name":         norm.get("last_name"),
                    "birthday":          norm.get("birthday"),
                    "contact_phone":     norm.get("contact_phone") or None,
                    "phone_type":        norm.get("phone_type"),
                    "contact_email":     norm.get("contact_email"),
                    "is_owner":          norm.get("is_owner", False),
                    "owner_type":        norm.get("owner_type", "unknown"),
                    "lead_type":         norm.get("lead_type", "unknown"),
                    "agency_name":       norm.get("agency_name"),
                    "discovery_source":  norm.get("discovery_source", "import"),
                    "contact_source":    "import",
                    "contact_confidence": norm.get("contact_confidence", 0),
                    "condition":         norm.get("condition"),
                    "sources_json":      norm.get("sources_json", "[]"),
                    "days_on_market":    norm.get("days_on_market", 0),
                    "first_seen_at":     datetime.utcnow(),
                    "last_seen_at":      datetime.utcnow(),
                    "crm_stage":         "novo",
                }

                with get_db() as db:
                    repo = LeadRepo(db)
                    _, created = repo.upsert(fp, lead_data)
                    if created:
                        stats.created += 1
                    else:
                        stats.updated += 1

            except Exception as e:
                stats.errors += 1
                log.debug("[import] row {i} error: {e}", i=i + 1, e=e)

    def _rescore(self) -> None:
        """Re-score all pending leads after import."""
        try:
            from scoring.scorer import Scorer
            n = Scorer().score_all_pending()
            log.info("[import] Re-scored {n} leads", n=n)
        except Exception as e:
            log.warning("[import] Scoring failed: {e}", e=e)

    # ── Export helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _export_csv(rows: list[dict], path: str) -> None:
        _EXPORT_COLS = [
            "title", "contact_name", "first_name", "last_name", "contact_phone",
            "phone_type", "contact_email", "birthday", "zone", "municipality",
            "parish", "typology", "property_type", "price", "area_m2",
            "is_owner", "owner_type", "lead_type", "discovery_source", "url",
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=_EXPORT_COLS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def _export_xlsx(rows: list[dict], path: str) -> None:
        from openpyxl import Workbook

        _EXPORT_COLS = [
            "title", "contact_name", "first_name", "last_name", "contact_phone",
            "phone_type", "contact_email", "birthday", "zone", "municipality",
            "parish", "typology", "property_type", "price", "area_m2",
            "is_owner", "owner_type", "lead_type", "discovery_source", "url",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Normalizados"
        ws.append(_EXPORT_COLS)
        for row in rows:
            ws.append([row.get(c, "") for c in _EXPORT_COLS])
        wb.save(path)
